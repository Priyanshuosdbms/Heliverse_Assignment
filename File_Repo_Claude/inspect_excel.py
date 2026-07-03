"""
tools/inspect_excel.py

Pure-python (no LLM) helper that gives the agent a *compact* view of a
register-map workbook: header rows + a sample of data rows, per sheet,
including the things that matter for grammar-sniffing on this file family:
  - the "row type" column (生成アイテム: Group / Reg)
  - merged-cell ranges (a Group row's name often spans merged cells)
  - fill color of each row (Group rows are often shaded)

This keeps token usage bounded regardless of how many thousand rows the
real workbook has — the LLM only ever sees a representative slice.
"""
from __future__ import annotations
import json
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def _cell_fill(cell) -> str | None:
    try:
        fg = cell.fill.fgColor
        if fg and fg.type == "rgb" and fg.rgb not in (None, "00000000"):
            return fg.rgb
    except Exception:
        pass
    return None


def inspect_workbook(xlsx_path: str, sample_rows: int = 40, sheets: list[str] | None = None) -> str:
    """Return a JSON string describing sheet names and a sample of rows
    (values + fill color) for each sheet, capped at `sample_rows` rows
    after the header so the LLM gets signal without a token explosion.

    If `sheets` is given, only those sheet names are inspected/parsed —
    use this to skip irrelevant tabs (e.g. a verification or changelog
    sheet) without relying on the LLM to guess which ones matter.
    """
    wb = load_workbook(xlsx_path, data_only=True)
    target_sheets = sheets if sheets else wb.sheetnames
    missing = [s for s in target_sheets if s not in wb.sheetnames]
    if missing:
        raise ValueError(f"sheet(s) not found in workbook: {missing}. Available: {wb.sheetnames}")

    report = {"file": Path(xlsx_path).name, "sheets": {}}

    for sheet_name in target_sheets:
        ws = wb[sheet_name]
        max_col = min(ws.max_column, 20)  # register sheets rarely need >20 cols
        merges = [str(m) for m in ws.merged_cells.ranges][:50]

        rows_out = []
        for r_idx, row in enumerate(
            ws.iter_rows(min_row=1, max_row=min(ws.max_row, sample_rows), max_col=max_col),
            start=1,
        ):
            row_vals = []
            row_fill = None
            for cell in row:
                row_vals.append(cell.value)
                fill = _cell_fill(cell)
                if fill:
                    row_fill = fill
            rows_out.append({"row": r_idx, "values": row_vals, "fill": row_fill})

        report["sheets"][sheet_name] = {
            "max_row": ws.max_row,
            "max_col": ws.max_column,
            "columns_sampled": max_col,
            "column_letters": [get_column_letter(i) for i in range(1, max_col + 1)],
            "merged_ranges_sample": merges,
            "rows": rows_out,
        }

    return json.dumps(report, ensure_ascii=False, default=str)


if __name__ == "__main__":
    import sys
    print(inspect_workbook(sys.argv[1]))